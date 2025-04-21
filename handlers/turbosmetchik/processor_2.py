# handlers/turbosmetchik/handler_v2.py
import openpyxl
import traceback
# Используем АБСОЛЮТНЫЙ импорт utils
from utils import is_likely_empty, check_merge, get_start_coord, is_integer_like

def process_turbosmetchik_2(input_path):
    """
    ОБРАБАТЫВАЕТ один Excel файл по логике "Турбосметчик-2".
    Отличается от v1 маппингом 5-й колонки выхода (Кол-во) на колонку N входа.
    ВОЗВРАЩАЕТ данные (заголовки и координаты) для дальнейшей обработки.

    Args:
        input_path (str): Путь к входному Excel файлу.

    Returns:
        tuple: Кортеж (output_headers, all_coords_data) или (None, None) в случае ошибки.
    """
    # print(f"\n--- Обработка файла (Турбосметчик-2): {os.path.basename(input_path)} ---")

    output_headers = ["№№ п/п", "Шифр расценки и коды ресурсов", "Наименование работ и затрат", "Единица измерения", "Кол-во единиц", "ВСЕГО затрат, руб."]
    start_id_col_idx = 0    # A

    processed_rows_list = []
    active_items_buffer = []
    pending_section_header = None
    pending_subsection_header = None
    first_section_found = False

    workbook = None
    try:
        workbook = openpyxl.load_workbook(filename=input_path, data_only=True)
        if not workbook.sheetnames: return None, None
        worksheet = workbook[workbook.sheetnames[0]]
        # print(f"Обработка листа '{worksheet.title}'...")

        for row_num, row_cells_tuple in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            row_cells = list(row_cells_tuple)
            non_empty_info = [(i, getattr(c, 'coordinate', 'N/A'), c.value) for i, c in enumerate(row_cells) if not is_likely_empty(c.value)]
            if not non_empty_info: continue

            cell_A = row_cells[start_id_col_idx] if len(row_cells) > start_id_col_idx else None
            cell_A_value_str = str(cell_A.value).strip() if cell_A and not is_likely_empty(cell_A.value) else ""

            row_type = None
            header_merge_coord = check_merge(worksheet, row_num, 0, 22) # A-W
            footer_merge_coord_DK = check_merge(worksheet, row_num, 3, 10) # D-K
            cell_D = row_cells[3] if len(row_cells) > 3 else None
            cell_D_value_str = str(cell_D.value).strip() if cell_D and not is_likely_empty(cell_D.value) else ""

            if header_merge_coord and cell_A_value_str.startswith("Раздел"): row_type = "section_header"
            elif header_merge_coord and cell_A_value_str.startswith("Подраздел"): row_type = "subsection_header"
            elif footer_merge_coord_DK and cell_D_value_str.startswith("Итого по подразделу"): row_type = "subsection_footer"
            elif footer_merge_coord_DK and cell_D_value_str.startswith("Итого по разделу"): row_type = "section_footer"
            else:
                dr_merge_coord = check_merge(worksheet, row_num, 3, 17) # D-R
                cell_D_price_text = row_cells[3] if len(row_cells) > 3 else None
                if dr_merge_coord and cell_D_price_text and str(cell_D_price_text.value).strip() == "Всего по позиции": row_type = "item_price_row"
                else:
                    if cell_A and cell_A.data_type != 'f' and not is_likely_empty(cell_A.value):
                         try:
                             float(str(cell_A.value).replace(',', '.').strip())
                             row_type = "item"
                         except (ValueError, TypeError): pass

            # --- Обработка буфера и состояний (аналогично v1) ---
            if row_type in ["section_header", "subsection_header", "section_footer", "subsection_footer"] and active_items_buffer:
                if first_section_found: processed_rows_list.extend(active_items_buffer)
                active_items_buffer = []

            # --- Обработка по типам строк (аналогично v1, кроме item) ---
            if row_type == "section_header":
                if first_section_found:
                    if pending_subsection_header: processed_rows_list.append(pending_subsection_header)
                    if pending_section_header: processed_rows_list.append(pending_section_header)
                pending_section_header = {"type": "header", "level": "section", "start_row": row_num, "col_1_coord": header_merge_coord, "col_3_value": cell_A_value_str, "col_6_coord": None}; pending_subsection_header = None; first_section_found = True;
            elif row_type == "subsection_header":
                if first_section_found and pending_subsection_header: processed_rows_list.append(pending_subsection_header)
                pending_subsection_header = {"type": "header", "level": "subsection", "start_row": row_num, "col_1_coord": header_merge_coord, "col_3_value": cell_A_value_str, "col_6_coord": None};
            elif row_type == "subsection_footer":
                 if pending_subsection_header:
                    cell_V_footer = row_cells[21] if len(row_cells) > 21 else None
                    footer_total_coord = cell_V_footer.coordinate if cell_V_footer else None
                    pending_subsection_header["col_6_coord"] = footer_total_coord
                    if first_section_found:
                        processed_rows_list.append(pending_subsection_header);
                        pending_subsection_header = None
            elif row_type == "section_footer":
                if first_section_found and pending_subsection_header: processed_rows_list.append(pending_subsection_header); pending_subsection_header = None
                if pending_section_header:
                    cell_V_footer = row_cells[21] if len(row_cells) > 21 else None
                    footer_total_coord = cell_V_footer.coordinate if cell_V_footer else None
                    pending_section_header["col_6_coord"] = footer_total_coord
                    if first_section_found:
                        processed_rows_list.append(pending_section_header);
                        pending_section_header = None
            elif row_type == "item_price_row":
                cell_V_price = row_cells[21] if len(row_cells) > 21 else None
                price_total_coord = cell_V_price.coordinate if cell_V_price else None
                for item in active_items_buffer:
                    item["col_6_coord"] = price_total_coord
                if first_section_found and active_items_buffer:
                     processed_rows_list.extend(active_items_buffer);
                     active_items_buffer = []
            elif row_type == "item":
                if first_section_found:
                    item_is_integer = is_integer_like(cell_A.value)
                    item_data = {"type": "item", "start_row": row_num, "col_6_coord": None}
                    # **** ИЗМЕНЕНИЕ ЗДЕСЬ ****
                    # Маппинг колонок входного файла (0-based index) на выходные колонки (1-based key)
                    # Выход: 1:"№№" 2:"Шифр" 3:"Наименование" 4:"Ед.изм." 5:"Кол-во"
                    # Вход T2: A(0)   B(1)    D(3)           L(11)       N(13) <-- ОТЛИЧИЕ от T1
                    input_indices_map = {1: 0, 2: 1, 3: 3, 4: 11, 5: 13}
                    # **** КОНЕЦ ИЗМЕНЕНИЯ ****

                    for out_col_num, in_col_idx in input_indices_map.items():
                        if in_col_idx < len(row_cells):
                            cell = row_cells[in_col_idx]
                            coord_str = getattr(cell, 'coordinate', 'N/A')
                            item_data[f"col_{out_col_num}_coord"] = coord_str
                        else:
                            item_data[f"col_{out_col_num}_coord"] = None

                    inline_price_coord = None
                    if item_is_integer:
                        merge_VW_coord = check_merge(worksheet, row_num, 21, 22) # V-W
                        if merge_VW_coord:
                            cell_V_inline = row_cells[21] if len(row_cells) > 21 else None
                            if cell_V_inline and not is_likely_empty(cell_V_inline.value):
                                inline_price_coord = get_start_coord(merge_VW_coord)

                    if inline_price_coord:
                        item_data["col_6_coord"] = inline_price_coord
                        processed_rows_list.append(item_data)
                    else:
                        active_items_buffer.append(item_data)

        # --- Обработка оставшихся данных (аналогично v1) ---
        if first_section_found:
            if active_items_buffer: processed_rows_list.extend(active_items_buffer)
            if pending_subsection_header: processed_rows_list.append(pending_subsection_header)
            if pending_section_header: processed_rows_list.append(pending_section_header)

        # --- ФОРМИРОВАНИЕ СПИСКА КООРДИНАТ (аналогично v1) ---
        processed_rows_list.sort(key=lambda x: x.get('start_row', float('inf')))
        all_coords_data = []
        for row_data in processed_rows_list:
            coords_row = [None] * len(output_headers)
            item_type = row_data.get("type")

            if item_type == "header":
                coords_row[0] = get_start_coord(row_data.get("col_1_coord"))
                coords_row[2] = row_data.get("col_3_value")
                coords_row[5] = get_start_coord(row_data.get("col_6_coord"))
            elif item_type == "item":
                for i_col in range(5):
                    coords_row[i_col] = get_start_coord(row_data.get(f"col_{i_col+1}_coord"))
                coords_row[5] = get_start_coord(row_data.get("col_6_coord"))

            all_coords_data.append(coords_row)

        return output_headers, all_coords_data

    except FileNotFoundError:
        print(f"[ОШИБКА] Файл не найден: {input_path}")
        return None, None
    except Exception as e:
        print(f"[КРИТИЧЕСКАЯ ОШИБКА] при обработке файла '{input_path}' (Турбосметчик-2): {e}")
        print("-" * 60); traceback.print_exc(); print("-" * 60)
        return None, None
    finally:
        if workbook:
            try: workbook.close()
            except: pass