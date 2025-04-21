# handlers/grandsmeta/processor.py
import openpyxl
import traceback
# Используем АБСОЛЮТНЫЙ импорт utils
from utils import is_likely_empty, check_merge, get_start_coord, is_integer_like

# Переименовываем функцию и обновляем docstring
def process_grandsmeta(input_path):
    """
    ОБРАБАТЫВАЕТ один Excel файл по логике "ГрандСМЕТА".
    ВОЗВРАЩАЕТ данные (заголовки и координаты) для дальнейшей обработки.

    Args:
        input_path (str): Путь к входному Excel файлу.

    Returns:
        tuple: Кортеж (output_headers, all_coords_data) или (None, None) в случае ошибки.
    """
    # print(f"\n--- Обработка файла (ГрандСМЕТА): {os.path.basename(input_path)} ---")

    output_headers = ["№№ п/п", "Шифр расценки и коды ресурсов", "Наименование работ и затрат", "Единица измерения", "Кол-во единиц", "ВСЕГО затрат, руб."]
    start_id_col_idx = 0    # A
    # Индексы колонок для новых правил
    col_A_idx = 0
    col_B_idx = 1
    col_C_idx = 2
    col_D_idx = 3
    col_E_idx = 4
    col_H_idx = 7
    col_K_idx = 10
    col_V_idx = 21 # Для итога по подразделу

    processed_rows_list = []
    active_items_buffer = []
    pending_section_header = None
    pending_subsection_header = None
    first_section_found = False # Оставляем флаг для отслеживания начала данных

    workbook = None
    try:
        workbook = openpyxl.load_workbook(filename=input_path, data_only=True)
        if not workbook.sheetnames:
            return None, None
        worksheet = workbook[workbook.sheetnames[0]]
        max_col_idx = worksheet.max_column - 1 # Безопасный доступ к колонкам

        for row_num, row_cells_tuple in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            row_cells = list(row_cells_tuple)
            # Пропускаем пустые строки
            non_empty_info = [(i, getattr(c, 'coordinate', 'N/A'), c.value) for i, c in enumerate(row_cells) if not is_likely_empty(c.value)]
            if not non_empty_info:
                continue

            # Получаем нужные ячейки безопасно
            cell_A = row_cells[col_A_idx] if col_A_idx <= max_col_idx else None
            cell_A_value_str = str(cell_A.value).strip() if cell_A and not is_likely_empty(cell_A.value) else ""

            cell_C = row_cells[col_C_idx] if col_C_idx <= max_col_idx else None
            cell_C_value_str = str(cell_C.value).strip() if cell_C and not is_likely_empty(cell_C.value) else ""

            cell_D = row_cells[col_D_idx] if col_D_idx <= max_col_idx else None
            cell_D_value_str = str(cell_D.value).strip() if cell_D and not is_likely_empty(cell_D.value) else ""

            row_type = None
            # --- НОВЫЕ ПРОВЕРКИ ТИПА СТРОКИ ---

            # 1. Заголовки (Раздел и Подраздел)
            header_merge_coord_AK = check_merge(worksheet, row_num, col_A_idx, col_K_idx) # A-K
            if header_merge_coord_AK:
                if cell_A_value_str.startswith("Раздел"):
                    row_type = "section_header"
                # Для подраздела по новым правилам проверяем только merge A-K
                else:
                    # Добавим проверку, чтобы не считать пустые заголовки подразделами
                    if cell_A_value_str:
                         row_type = "subsection_header"


            # 2. Футеры (Раздел и Подраздел)
            if not row_type:
                footer_merge_coord_DK = check_merge(worksheet, row_num, col_D_idx, col_K_idx) # D-K
                if footer_merge_coord_DK and cell_D_value_str.startswith("Итого по подразделу"):
                    row_type = "subsection_footer"
                else:
                    footer_merge_coord_CH = check_merge(worksheet, row_num, col_C_idx, col_H_idx) # C-H
                    if footer_merge_coord_CH and cell_C_value_str.startswith("Итого по разделу"):
                        row_type = "section_footer"

            # 3. Строка "Всего по позиции"
            if not row_type:
                 # По новым правилам проверяем только текст в C, без merge
                 if cell_C_value_str == "Всего по позиции":
                     row_type = "item_price_row"

            # 4. Строка Item
            if not row_type:
                # Проверка на item осталась прежней (число в колонке A)
                if cell_A and cell_A.data_type != 'f' and not is_likely_empty(cell_A.value):
                     try:
                         float(str(cell_A.value).replace(',', '.').strip())
                         row_type = "item"
                     except (ValueError, TypeError):
                         pass # Не число, не item

            # --- Обработка буфера и состояний (логика осталась прежней) ---
            if row_type in ["section_header", "subsection_header", "section_footer", "subsection_footer"] and active_items_buffer:
                if first_section_found:
                    processed_rows_list.extend(active_items_buffer)
                active_items_buffer = []

            # --- Обработка по типам строк (НОВАЯ ЛОГИКА) ---
            if row_type == "section_header":
                if first_section_found:
                    # Сначала обрабатываем незавершенные предыдущие заголовки
                    if pending_subsection_header: processed_rows_list.append(pending_subsection_header)
                    if pending_section_header: processed_rows_list.append(pending_section_header)
                # Сохраняем новый заголовок раздела
                pending_section_header = {"type": "header", "level": "section", "start_row": row_num, "col_1_coord": header_merge_coord_AK, "col_3_value": cell_A_value_str, "col_6_coord": None}
                pending_subsection_header = None # Сбрасываем подраздел
                first_section_found = True
            elif row_type == "subsection_header":
                 # Обрабатываем предыдущий подраздел, если был
                if first_section_found and pending_subsection_header: processed_rows_list.append(pending_subsection_header)
                 # Сохраняем новый заголовок подраздела
                pending_subsection_header = {"type": "header", "level": "subsection", "start_row": row_num, "col_1_coord": header_merge_coord_AK, "col_3_value": cell_A_value_str, "col_6_coord": None}
            elif row_type == "subsection_footer":
                 if pending_subsection_header:
                    # Итоговая сумма для футера подраздела берется из колонки V (индекс 21)
                    cell_V_footer = row_cells[col_V_idx] if col_V_idx <= max_col_idx else None
                    footer_total_coord = cell_V_footer.coordinate if cell_V_footer else None
                    pending_subsection_header["col_6_coord"] = footer_total_coord # Координата итога (V)
                    if first_section_found:
                        processed_rows_list.append(pending_subsection_header)
                        pending_subsection_header = None # Очищаем, т.к. подраздел завершен
            elif row_type == "section_footer":
                # Сначала обрабатываем незавершенный подраздел, если он был перед итогом раздела
                if first_section_found and pending_subsection_header:
                    processed_rows_list.append(pending_subsection_header)
                    pending_subsection_header = None
                if pending_section_header:
                    # Итоговая сумма для футера раздела берется из колонки K (индекс 10)
                    cell_K_footer = row_cells[col_K_idx] if col_K_idx <= max_col_idx else None
                    footer_total_coord = cell_K_footer.coordinate if cell_K_footer else None
                    pending_section_header["col_6_coord"] = footer_total_coord # Координата итога (K)
                    if first_section_found:
                        processed_rows_list.append(pending_section_header)
                        pending_section_header = None # Очищаем, т.к. раздел завершен
            elif row_type == "item_price_row":
                # Цена для item'ов в буфере берется из колонки K (индекс 10) строки "Всего по позиции"
                cell_K_price = row_cells[col_K_idx] if col_K_idx <= max_col_idx else None
                price_total_coord = cell_K_price.coordinate if cell_K_price else None
                for item in active_items_buffer:
                    item["col_6_coord"] = price_total_coord
                if first_section_found and active_items_buffer:
                     processed_rows_list.extend(active_items_buffer)
                     active_items_buffer = [] # Очищаем буфер после присвоения цены
            elif row_type == "item":
                if first_section_found:
                    item_is_integer = is_integer_like(cell_A.value)
                    item_data = {"type": "item", "start_row": row_num, "col_6_coord": None}

                    # **** НОВЫЙ МАППИНГ КОЛОНОК для ГрандСМЕТА ****
                    # Выход: 1:"№№" 2:"Шифр" 3:"Наименование" 4:"Ед.изм." 5:"Кол-во"
                    # Вход GS: A(0)   B(1)    C(2)           D(3)       E(4)
                    input_indices_map = {1: col_A_idx, 2: col_B_idx, 3: col_C_idx, 4: col_D_idx, 5: col_E_idx}

                    for out_col_num, in_col_idx in input_indices_map.items():
                        if in_col_idx <= max_col_idx:
                            cell = row_cells[in_col_idx]
                            coord_str = getattr(cell, 'coordinate', 'N/A') # Получаем координату
                            item_data[f"col_{out_col_num}_coord"] = coord_str # Сохраняем координату
                        else:
                            item_data[f"col_{out_col_num}_coord"] = None # Если колонки нет

                    inline_price_coord = None
                    # --- НОВАЯ ПРОВЕРКА ВСТРОЕННОЙ ЦЕНЫ ---
                    # Правило: Целое число в A И НЕ пустая ячейка K
                    if item_is_integer:
                        cell_K_inline = row_cells[col_K_idx] if col_K_idx <= max_col_idx else None
                        if cell_K_inline and not is_likely_empty(cell_K_inline.value):
                             # Цена находится в ячейке K (индекс 10), без проверки merge
                            inline_price_coord = cell_K_inline.coordinate

                    if inline_price_coord:
                        # Если цена найдена в строке (в K), присваиваем ее координату и добавляем item сразу
                        item_data["col_6_coord"] = inline_price_coord
                        processed_rows_list.append(item_data)
                    else:
                        # Если цена не найдена (или номер не целый, или K пустое), добавляем item в буфер
                        active_items_buffer.append(item_data)

        # --- Обработка оставшихся данных в конце файла (логика осталась) ---
        if first_section_found:
            if active_items_buffer: processed_rows_list.extend(active_items_buffer)
            if pending_subsection_header: processed_rows_list.append(pending_subsection_header)
            if pending_section_header: processed_rows_list.append(pending_section_header)

        # --- ФОРМИРОВАНИЕ СПИСКА КООРДИНАТ (обновляем под новые координаты) ---
        processed_rows_list.sort(key=lambda x: x.get('start_row', float('inf')))
        all_coords_data = []
        for row_data in processed_rows_list:
            coords_row = [None] * len(output_headers)
            item_type = row_data.get("type")

            if item_type == "header":
                # Координата A-K для №№
                coords_row[0] = get_start_coord(row_data.get("col_1_coord"))
                # Текст заголовка для Наименования
                coords_row[2] = row_data.get("col_3_value")
                # Координата итога (V для подраздела, K для раздела) для ВСЕГО
                coords_row[5] = get_start_coord(row_data.get("col_6_coord"))
            elif item_type == "item":
                # Копируем координаты для колонок 1-5 из сохраненных данных (A, B, C, D, E)
                for i_col in range(5): # 0 to 4
                    coords_row[i_col] = get_start_coord(row_data.get(f"col_{i_col+1}_coord"))
                # Координата для колонки 6 (ВСЕГО) - из K (inline) или K (из "Всего по позиции")
                coords_row[5] = get_start_coord(row_data.get("col_6_coord"))

            all_coords_data.append(coords_row)

        return output_headers, all_coords_data

    except FileNotFoundError:
        print(f"[ОШИБКА] Файл не найден: {input_path}")
        return None, None
    except Exception as e:
        # Обновляем сообщение об ошибке
        print(f"[КРИТИЧЕСКАЯ ОШИБКА] при обработке файла '{input_path}' (ГрандСМЕТА): {e}")
        print("-" * 60); traceback.print_exc(); print("-" * 60)
        return None, None
    finally:
        if workbook:
            try: workbook.close()
            except: pass