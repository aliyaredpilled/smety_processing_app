# handlers/smeta_ru/handler.py
import openpyxl
import traceback
# Используем АБСОЛЮТНЫЙ импорт для доступа к utils.py из корневой папки
from utils import is_likely_empty, check_merge, get_start_coord, is_zero

def process_smeta_ru(input_path):
    """
    ОБРАБАТЫВАЕТ один Excel файл по логике "Смета ру".
    НЕ СОХРАНЯЕТ ФАЙЛ, а ВОЗВРАЩАЕТ данные для дальнейшей обработки.

    Args:
        input_path (str): Путь к входному Excel файлу.

    Returns:
        tuple: Кортеж (output_headers, all_coords_data) или (None, None) в случае ошибки.
               output_headers (list): Список заголовков для выходного файла.
               all_coords_data (list): Список списков с данными (строки координат).
    """
    # print(f"\n--- Обработка файла (Смета ру): {os.path.basename(input_path)} ---") # Убрал print для чистоты вывода в app

    # --- Конфигурация ---
    output_headers = ["№№ п/п", "Шифр расценки и коды ресурсов", "Наименование работ и затрат", "Единица измерения", "Кол-во единиц", "ВСЕГО затрат, руб."]
    # Индексы колонок (0-based)
    price_total_col_idx = 8 # I (Итоговая цена)
    item_individual_price_col_idx = 9 # J (Индивидуальная цена ресурса)
    price_ztr_col_idx = 10  # K (Стоимость единицы)
    start_id_col_idx = 0    # A (Номер п/п или начало шифра)

    # Условия для определения строки с ЦЕНОЙ ИТОГО по позиции (price_row)
    # Должны быть НЕ ПУСТЫМИ колонки I и K
    price_row_must_be_non_empty_indices = {price_total_col_idx, price_ztr_col_idx} # {8, 10}
    # Должны быть ПУСТЫМИ колонки A-H
    price_row_must_be_empty_indices = set(range(8)) # {0, 1, 2, 3, 4, 5, 6, 7}

    # --- Инициализация ---
    processed_rows_list = [] # Список для хранения словарей обработанных строк (item, header)
    active_items_buffer = [] # Буфер для item'ов, ожидающих свою цену
    pending_section_header = None # Данные о текущем незакрытом разделе
    pending_subsection_header = None # Данные о текущем незакрытом подразделе
    first_section_found = False # Флаг, что мы начали обрабатывать данные внутри первого раздела
    skipped_items_zero_j_count = 0 # Счетчик пропущенных item'ов с нулевой ценой в J

    workbook = None # Инициализация для блока finally

    try:
        # --- Загрузка Excel ---
        # print(f"Загрузка: {input_path}")
        workbook = openpyxl.load_workbook(filename=input_path, data_only=True) # data_only=True для чтения значений, а не формул
        if not workbook.sheetnames:
            print(f"Ошибка: Нет листов в файле '{input_path}'.")
            return None, None
        worksheet = workbook[workbook.sheetnames[0]] # Берем первый лист
        # print(f"Обработка листа '{worksheet.title}'...")

        # --- ПАРСИНГ ДАННЫХ ---
        # Итерация по строкам, начиная со второй (пропускаем заголовки)
        for row_num, row_cells_tuple in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
            row_cells = list(row_cells_tuple) # Преобразуем кортеж в список для удобства

            # Собираем информацию о непустых ячейках для анализа строки
            non_empty_info = [(idx, cell.coordinate, cell.value)
                              for idx, cell in enumerate(row_cells)
                              if not is_likely_empty(cell.value)]

            # Если вся строка пустая, пропускаем её
            if not non_empty_info:
                continue

            # Получаем ключевые ячейки для определения типа строки
            cell_A = row_cells[start_id_col_idx] if len(row_cells) > start_id_col_idx else None
            cell_I = row_cells[price_total_col_idx] if len(row_cells) > price_total_col_idx else None
            cell_J = row_cells[item_individual_price_col_idx] if len(row_cells) > item_individual_price_col_idx else None

            # Получаем строковое представление значения ячейки A (для проверки заголовков)
            cell_A_value_str = str(cell_A.value).strip() if cell_A and not is_likely_empty(cell_A.value) else ""

            # --- Определение типа строки ---
            row_type = None
            # Проверяем наличие merge для заголовков/футеров (Смета.ру использует merge A-K для заголовков/футеров разделов)
            # Широкий диапазон для заголовков/футеров
            header_footer_merge_coord_AK = check_merge(worksheet, row_num, 0, 10) # A-K

            # Проверка на Заголовок Раздела/Подраздела
            if header_footer_merge_coord_AK:
                 if cell_A_value_str.startswith("Раздел:"):
                     row_type = "section_header"
                 elif cell_A_value_str.startswith("Подраздел:"):
                     row_type = "subsection_header"
                 # Проверка на Футер Раздела/Подраздела (Итого по...)
                 elif cell_A_value_str.startswith("Итого по подразделу:"):
                     row_type = "subsection_footer"
                 elif cell_A_value_str.startswith("Итого по разделу:"):
                     row_type = "section_footer"

            # Если это не заголовок/футер, проверяем другие типы
            if row_type is None:
                # Получаем множество индексов непустых ячеек
                non_empty_cell_indices = {info[0] for info in non_empty_info}

                # Условие для строки с итоговой ценой (price_row):
                # 1. Колонки I и K (must_be_non_empty) НЕ ПУСТЫЕ
                cond1 = non_empty_cell_indices.issuperset(price_row_must_be_non_empty_indices)
                # 2. Колонки A-H (must_be_empty) ПУСТЫЕ (пересечение множеств пустое)
                cond2 = non_empty_cell_indices.isdisjoint(price_row_must_be_empty_indices)

                if cond1 and cond2:
                    row_type = "item_price_row"
                else:
                    # Условие для строки с позицией (item):
                    # 1. Ячейка A не пустая
                    # 2. Ячейка A не формула (проверяем data_type)
                    # 3. Значение в A похоже на число (номер п/п или начало шифра)
                    if cell_A and cell_A.data_type != 'f' and not is_likely_empty(cell_A.value):
                         # Пытаемся преобразовать значение в float, чтобы убедиться, что это число или числовой код
                         try:
                             float(str(cell_A.value).replace(',', '.').strip())
                             row_type = "item"
                         except (ValueError, TypeError):
                             # Если не удалось преобразовать, это не стандартная строка item (возможно, текст или подзаголовок без merge)
                             # print(f"  [DEBUG] Строка {row_num}: A='{cell_A.value}' не похожа на числовой item.")
                             pass # row_type остается None

            # --- Логика обработки в зависимости от типа строки ---

            # Если начинается новый раздел/подраздел или заканчивается старый,
            # а в буфере есть item'ы без цены, то добавляем их "как есть" (с пустой ценой)
            if row_type in ["section_header", "subsection_header", "section_footer", "subsection_footer"] and active_items_buffer:
                if first_section_found: # Добавляем только если уже был найден первый раздел
                    # print(f"  [WARN] Строка {row_num} ({row_type}): Найдены элементы ({len(active_items_buffer)}) без цены перед заголовком/футером. Добавляем.")
                    processed_rows_list.extend(active_items_buffer)
                # else:
                    # print(f"  [DEBUG] Строка {row_num} ({row_type}): Найдены элементы ({len(active_items_buffer)}) до первого раздела. Очищаем буфер.")
                active_items_buffer = [] # Очищаем буфер в любом случае

            # Обработка Заголовка Раздела
            if row_type == "section_header":
                if first_section_found: # Если это не первый раздел
                    # Проверяем, остались ли незакрытые предыдущие раздел/подраздел
                    if pending_subsection_header:
                        # print(f"  [WARN] Строка {row_num} (section_header): Найден незакрытый ПОДРАЗДЕЛ {pending_subsection_header['start_row']}. Добавляем.")
                        processed_rows_list.append(pending_subsection_header)
                    if pending_section_header:
                        # print(f"  [WARN] Строка {row_num} (section_header): Найден незакрытый РАЗДЕЛ {pending_section_header['start_row']}. Добавляем.")
                        processed_rows_list.append(pending_section_header)

                # Сохраняем данные нового раздела
                pending_section_header = {
                    "type": "header",
                    "level": "section",
                    "start_row": row_num,
                    "col_1_coord": header_footer_merge_coord_AK, # Координата объединенной ячейки A-K
                    "col_3_value": cell_A_value_str,            # Текст заголовка (из A)
                    "col_6_value": None,                        # Итоговая цена (пока неизвестна)
                    "col_6_coord": None                         # Координата итог. цены (пока неизвестна)
                }
                pending_subsection_header = None # Сбрасываем активный подраздел
                first_section_found = True # Отмечаем, что нашли первый раздел
                # print(f"Найден раздел: строка {row_num}")

            # Обработка Заголовка Подраздела
            elif row_type == "subsection_header":
                # Если был предыдущий незакрытый подраздел, добавляем его
                if first_section_found and pending_subsection_header:
                    # print(f"  [WARN] Строка {row_num} (subsection_header): Найден незакрытый ПОДРАЗДЕЛ {pending_subsection_header['start_row']}. Добавляем.")
                    processed_rows_list.append(pending_subsection_header)

                # Сохраняем данные нового подраздела
                pending_subsection_header = {
                    "type": "header",
                    "level": "subsection",
                    "start_row": row_num,
                    "col_1_coord": header_footer_merge_coord_AK, # Координата объединенной ячейки A-K
                    "col_3_value": cell_A_value_str,            # Текст заголовка (из A)
                    "col_6_value": None,                        # Итоговая цена (пока неизвестна)
                    "col_6_coord": None                         # Координата итог. цены (пока неизвестна)
                }
                # print(f"Найден подраздел: строка {row_num}")

            # Обработка Футера Подраздела ("Итого по подразделу")
            elif row_type == "subsection_footer":
                if pending_subsection_header:
                    # Обновляем данные активного подраздела ценой из футера (колонка I)
                    pending_subsection_header["col_6_value"] = cell_I.value if cell_I else None
                    # Координата цены футера - это та же объединенная ячейка A-K
                    pending_subsection_header["col_6_coord"] = header_footer_merge_coord_AK
                    if first_section_found:
                        processed_rows_list.append(pending_subsection_header)
                        # print(f"Закрыт подраздел ({pending_subsection_header['start_row']}) футером {row_num}. Добавлен в processed_rows_list.")
                        pending_subsection_header = None # Сбрасываем активный подраздел
                    # else: # Этот случай маловероятен, если first_section_found проверяется раньше
                    #     print(f"  [WARN] Строка {row_num}: Найден итог подраздела, но первый раздел еще не найден.")
                elif first_section_found: # Если есть футер, но нет активного подраздела
                    print(f"  [WARN] Строка {row_num}: Итого по подразделу найдено, но не было активного подраздела для закрытия.")

            # Обработка Футера Раздела ("Итого по разделу")
            elif row_type == "section_footer":
                # Если остался незакрытый подраздел внутри этого раздела, добавляем его
                if first_section_found and pending_subsection_header:
                    # print(f"  [WARN] Строка {row_num} (section_footer): Найден незакрытый ПОДРАЗДЕЛ {pending_subsection_header['start_row']} перед итогом раздела. Добавляем.")
                    processed_rows_list.append(pending_subsection_header)
                    pending_subsection_header = None

                if pending_section_header:
                    # Обновляем данные активного раздела ценой из футера (колонка I)
                    pending_section_header["col_6_value"] = cell_I.value if cell_I else None
                    # Координата цены футера - это та же объединенная ячейка A-K
                    pending_section_header["col_6_coord"] = header_footer_merge_coord_AK
                    if first_section_found:
                        processed_rows_list.append(pending_section_header)
                        # print(f"Закрыт раздел ({pending_section_header['start_row']}) футером {row_num}. Добавлен в processed_rows_list.")
                        pending_section_header = None # Сбрасываем активный раздел
                    # else: # Маловероятно
                    #     print(f"  [WARN] Строка {row_num}: Найден итог раздела, но первый раздел еще не найден.")
                elif first_section_found: # Если есть футер, но нет активного раздела
                    print(f"  [WARN] Строка {row_num}: Итого по разделу найдено, но не было активного раздела для закрытия.")

            # Обработка Строки с Итоговой Ценой Позиции
            elif row_type == "item_price_row":
                price_total_value = cell_I.value if cell_I else None
                price_total_coord = cell_I.coordinate if cell_I else None
                # price_ztr_value = row_cells[price_ztr_col_idx].value if len(row_cells) > price_ztr_col_idx else None

                if active_items_buffer:
                    # print(f"  Строка {row_num}: Найдена цена ({price_total_value}) для {len(active_items_buffer)} элементов в буфере.")
                    # Присваиваем эту цену всем item'ам в буфере
                    for item in active_items_buffer:
                        item["col_6_value"] = price_total_value # Значение из I
                        item["col_6_coord"] = price_total_coord # Координата ячейки I
                    # Добавляем обработанные item'ы в общий список
                    if first_section_found:
                        processed_rows_list.extend(active_items_buffer)
                    active_items_buffer = [] # Очищаем буфер
                # else:
                    # print(f"  [DEBUG] Строка {row_num}: Найдена строка цены, но буфер item'ов пуст.")

            # Обработка Строки с Позицией (Item)
            elif row_type == "item":
                if first_section_found: # Обрабатываем item только если мы внутри раздела
                    # Проверка на нулевую цену ресурса в колонке J (по ТЗ)
                    cell_J_value = cell_J.value if cell_J else None
                    if is_zero(cell_J_value):
                        # print(f"  [Фильтр] Пропуск item {row_num} (A='{cell_A.value}'), т.к. значение в колонке J ({item_individual_price_col_idx+1}) равно 0.")
                        skipped_items_zero_j_count += 1
                    else:
                        # Собираем данные для item'а (координаты ячеек A-E)
                        item_data = {
                            "type": "item",
                            "start_row": row_num,
                            "col_6_value": None, # Цена пока неизвестна
                            "col_6_coord": None  # Координата цены пока неизвестна
                        }
                        # Заполняем координаты для первых 5 колонок вывода (из колонок A-E входа)
                        for i in range(min(5, len(row_cells))): # Берем не более 5 колонок
                            cell = row_cells[i]
                            # Сохраняем и значение, и координату
                            item_data[f"col_{i+1}_value"] = cell.value
                            item_data[f"col_{i+1}_coord"] = cell.coordinate
                        # Добавляем item в буфер ожидания цены
                        active_items_buffer.append(item_data)
                # else:
                #     print(f"  [DEBUG] Строка {row_num}: Найден item, но первый раздел еще не найден. Игнорируется.")

            # else: # Если row_type остался None
                # print(f"  [DEBUG] Строка {row_num}: Не удалось определить тип строки. Пропускается. A='{cell_A_value_str}', Непустые: {non_empty_info}")
                # pass # Просто пропускаем неопознанные строки

        # --- Обработка данных в конце файла ---
        if first_section_found:
            # Если остались item'ы в буфере без цены
            if active_items_buffer:
                # print(f"  [WARN] Конец файла: {len(active_items_buffer)} элементов остались в буфере без цены. Добавляем.")
                processed_rows_list.extend(active_items_buffer)
            # Если остался незакрытый подраздел
            if pending_subsection_header:
                # print(f"  [WARN] Конец файла: Найден незакрытый ПОДРАЗДЕЛ {pending_subsection_header['start_row']}. Добавляем.")
                processed_rows_list.append(pending_subsection_header)
            # Если остался незакрытый раздел
            if pending_section_header:
                # print(f"  [WARN] Конец файла: Найден незакрытый РАЗДЕЛ {pending_section_header['start_row']}. Добавляем.")
                processed_rows_list.append(pending_section_header)
        # else:
            # print("  [WARN] Первый раздел так и не был найден в файле.")

        # --- ФОРМИРОВАНИЕ СПИСКА ДАННЫХ КООРДИНАТ (all_coords_data) ---
        # print("\nФормирование списка данных координат...")
        # Сортируем собранные данные по номеру строки для правильного порядка
        processed_rows_list.sort(key=lambda x: x.get('start_row', float('inf')))

        all_coords_data = [] # Итоговый список списков с координатами
        skipped_final_price_count = 0

        for row_data in processed_rows_list:
            # Пропускаем строки с нулевой итоговой ценой (col_6_value) - новое требование
            total_cost_value = row_data.get("col_6_value")
            if is_zero(total_cost_value):
                # print(f"  [Фильтр Итог] Пропуск строки {row_data.get('start_row', 'N/A')} (тип: {row_data.get('type', 'N/A')}), итоговая цена = 0.")
                skipped_final_price_count += 1
                continue # Пропускаем эту запись

            # Создаем строку для выходного файла (пока пустую)
            coords_row = [None] * len(output_headers)
            item_type = row_data.get("type")

            if item_type == "header":
                # Для заголовков/футеров:
                # Колонка 1 (№№) -> Координата объединенной ячейки A-K
                coords_row[0] = get_start_coord(row_data.get("col_1_coord"))
                # Колонка 3 (Наименование) -> Текст заголовка/футера
                coords_row[2] = row_data.get("col_3_value")
                # Колонка 6 (ВСЕГО) -> Координата объединенной ячейки A-K (где лежит итоговая сумма)
                coords_row[5] = get_start_coord(row_data.get("col_6_coord"))
            elif item_type == "item":
                # Для позиций (item):
                # Колонки 1-5 -> Координаты ячеек A-E соответственно
                for i in range(min(5, len(output_headers))): # От 0 до 4
                    col_key_coord = f"col_{i+1}_coord"
                    coords_row[i] = get_start_coord(row_data.get(col_key_coord))
                # Колонка 6 (ВСЕГО) -> Координата ячейки с итоговой ценой (из колонки I исходного файла)
                if len(output_headers) > 5: # Индекс 5
                     coords_row[5] = get_start_coord(row_data.get("col_6_coord"))

            all_coords_data.append(coords_row)

        # print(f"Формирование данных завершено. Строк для записи: {len(all_coords_data)}.")
        # print(f"  Пропущено позиций с нулевой ценой в колонке J: {skipped_items_zero_j_count}")
        # print(f"  Пропущено записей с нулевой итоговой ценой (col_6): {skipped_final_price_count}")

        # --- ВОЗВРАЩАЕМ ЗАГОЛОВКИ И ДАННЫЕ ---
        return output_headers, all_coords_data

    except FileNotFoundError:
        print(f"[ОШИБКА] Файл не найден: {input_path}")
        return None, None
    except Exception as e:
        print(f"[КРИТИЧЕСКАЯ ОШИБКА] при обработке файла '{input_path}' (Смета ру): {e}")
        print("-" * 60)
        traceback.print_exc()
        print("-" * 60)
        return None, None
    finally:
        # Гарантированно закрываем workbook, если он был открыт
        if workbook:
            try:
                workbook.close()
            except Exception as close_e:
                 print(f"  [WARN] Не удалось закрыть Excel файл '{input_path}': {close_e}")