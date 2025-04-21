import os
import uuid
import zipfile
import shutil
import traceback
import time # Добавим для имитации работы и очистки
import openpyxl # <--- ДОБАВЛЕН ИМПОРТ OPENPYXL
import dispatcher # <<< ИМПОРТИРУЕМ НОВЫЙ МОДУЛЬ ДИСПЕТЧЕРА
import re # <<< ДОБАВЛЯЕМ ИМПОРТ RE
from flask import Flask, request, render_template, jsonify, send_from_directory, url_for
from werkzeug.utils import secure_filename
from openpyxl.styles import Alignment, Font # Убедитесь, что Font тоже импортирован
from formatting import (
    read_reference_widths,
    auto_adjust_column_width,
    apply_reference_widths,
    apply_formatting
)

# --- Конфигурация Flask ---
app = Flask(__name__)

# <<< НАЧАЛО ИЗМЕНЕНИЙ: Настройки для отключения кэша в разработке >>>
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0 # Отключает кэширование статических файлов браузером
app.config['TEMPLATES_AUTO_RELOAD'] = True # Автоматически перезагружает шаблоны при изменении
app.config['DEBUG'] = True # Убедимся, что включен режим отладки (он тоже влияет на кэш)
# <<< КОНЕЦ ИЗМЕНЕНИЙ >>>

UPLOAD_FOLDER = 'uploads'
RESULTS_FOLDER = 'results'
REFERENCE_FOLDER = 'reference_files'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(RESULTS_FOLDER, exist_ok=True)
os.makedirs(REFERENCE_FOLDER, exist_ok=True)

REFERENCE_SMETA_RU = os.path.join(REFERENCE_FOLDER, "Смета ру.xlsm") # Укажите точное имя вашего референсного файла
REFERENCE_TURBOSMETCHIK = os.path.join(REFERENCE_FOLDER, "Турбосметчик1,2,3.xlsm") # <-- ДОБАВЛЕНО

ALLOWED_EXTENSIONS = {'xlsx', 'xlsm', 'zip'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['RESULTS_FOLDER'] = RESULTS_FOLDER
# Ограничение размера файла (например, 100 MB) - раскомментируйте, если нужно
# app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024

# === Глобальный словарь для хранения статуса обработки ===
# Ключ - session_id, Значение - словарь {"processed": N, "total": M, "status": "...", "error": None}
processing_status = {}
# =========================================================

def allowed_file(filename):
    """Проверяет, имеет ли файл разрешенное расширение."""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# --- Маршруты ---

@app.route('/')
def index():
    """Отображает главную страницу с формой загрузки."""
    try:
        # Получаем доступные типы смет из процессора
        smeta_types = dispatcher.get_available_processor_types()
        return render_template('index.html', smeta_types=smeta_types)
    except Exception as e:
        print(f"Ошибка при загрузке типов смет: {e}")
        traceback.print_exc() # Добавим вывод traceback для диагностики
        # Можно вернуть страницу с ошибкой или пустой список
        return render_template('index.html', smeta_types=[], error="Не удалось загрузить типы смет.")


@app.route('/upload', methods=['POST'])
def upload_file():
    """Обрабатывает загрузку файла, запускает процессор и сохраняет статус."""
    # === Получаем ID сессии от клиента ===
    client_session_id = request.form.get('client_session_id')
    if not client_session_id:
        return jsonify({"success": False, "error": "Отсутствует ID сессии клиента"}), 400
    # =======================================

    if 'file' not in request.files: return jsonify({"success": False, "error": "Файл не выбран"}), 400
    file = request.files['file']
    smeta_type = request.form.get('smeta_type')
    if not smeta_type: return jsonify({"success": False, "error": "Тип сметы не выбран"}), 400
    if file.filename == '': return jsonify({"success": False, "error": "Файл не выбран"}), 400

    # Временная папка для конкретной сессии загрузки
    upload_path = os.path.join(app.config['UPLOAD_FOLDER'], client_session_id)
    # Убедимся, что папка чистая перед началом
    if os.path.exists(upload_path): shutil.rmtree(upload_path)
    os.makedirs(upload_path, exist_ok=True)

    saved_file_path = None # Определим позже

    try: # Главный try для всей обработки запроса
        if file and allowed_file(file.filename):
            original_filename_unsafe = file.filename
            print(f"({client_session_id}) Получен файл: {original_filename_unsafe}")
            is_zip = original_filename_unsafe.lower().endswith('.zip')
            is_excel = original_filename_unsafe.lower().endswith(('.xlsx', '.xlsm'))
            safe_filename_for_saving = secure_filename(original_filename_unsafe)
            if not safe_filename_for_saving:
                 extension = ".zip" if is_zip else (".xlsx" if is_excel else ".file")
                 safe_filename_for_saving = f"uploaded_file{extension}" # Упростили имя
                 print(f"({client_session_id}) secure_filename пустое, сгенерировано: {safe_filename_for_saving}")

            saved_file_path = os.path.join(upload_path, safe_filename_for_saving)
            file.save(saved_file_path)
            print(f"({client_session_id}) Сохранен как: {saved_file_path}")

            files_to_process_info = []
            if is_zip:
                # --- Инициализация статуса для ZIP ---
                processing_status[client_session_id] = {"processed": 0, "total": 0, "status": "Распаковка...", "error": None}
                print(f"({client_session_id}) ZIP архив. Распаковка...")
                # ------------------------------------
                extract_path = os.path.join(upload_path, 'extracted')
                os.makedirs(extract_path, exist_ok=True)
                try:
                    with zipfile.ZipFile(saved_file_path, 'r') as zip_ref:
                        # === Получаем и СОРТИРУЕМ список файлов ===
                        file_list_in_zip = sorted([m for m in zip_ref.infolist() if not m.is_dir() \
                                                  and not m.filename.startswith('__MACOSX/') \
                                                  and not m.filename.startswith('.') \
                                                  and m.filename.lower().endswith(('.xlsx', '.xlsm'))],
                                                  key=lambda member: member.filename)
                        # --- Обновляем total в статусе ---
                        processing_status[client_session_id]["total"] = len(file_list_in_zip)
                        processing_status[client_session_id]["status"] = "Найдено файлов: {}".format(len(file_list_in_zip))
                        print(f"({client_session_id}) Найдено и отсортировано файлов в ZIP: {[m.filename for m in file_list_in_zip]}")
                        # ---------------------------------

                        for member in file_list_in_zip:
                            try: filename_decoded = member.filename.encode('cp437').decode('utf-8', 'ignore')
                            except: filename_decoded = member.filename.encode('cp437').decode('cp437', 'ignore')
                            target_path = os.path.join(extract_path, os.path.basename(filename_decoded))
                            try:
                                with zip_ref.open(member) as source, open(target_path, "wb") as target: shutil.copyfileobj(source, target)
                                files_to_process_info.append({"path": target_path, "original_name": filename_decoded})
                            except Exception as extract_err: print(f"  [WARN] ({client_session_id}) Не удалось извлечь {filename_decoded}: {extract_err}")

                    if not files_to_process_info: raise ValueError("В архиве не найдено поддерживаемых Excel файлов.")

                except zipfile.BadZipFile: raise ValueError("Некорректный ZIP архив.")
                except ValueError as ve: raise ve
                except Exception as e: raise ValueError(f"Ошибка при распаковке ZIP: {e}")

            elif is_excel:
                print(f"({client_session_id}) Одиночный Excel файл.")
                 # --- Инициализация статуса для одного файла ---
                processing_status[client_session_id] = {"processed": 0, "total": 1, "status": "Подготовка...", "error": None}
                # ------------------------------------------
                files_to_process_info.append({"path": saved_file_path, "original_name": original_filename_unsafe})
            else:
                raise ValueError(f"Неподдерживаемый тип файла: {original_filename_unsafe}.")

            # --- Сбор данных от процессора и обновление статуса ---
            collected_results = []
            has_errors = False
            common_headers = None

            for i, file_info in enumerate(files_to_process_info):
                input_fpath = file_info["path"]
                original_fname = file_info["original_name"]

                # --- Обновляем статус перед обработкой файла ---
                current_total = processing_status.get(client_session_id, {}).get("total", len(files_to_process_info))
                processing_status[client_session_id]["status"] = f"Обработка файла {i+1} из {current_total}: {original_fname}..."
                print(f"\n({client_session_id}) {processing_status[client_session_id]['status']}")
                # ---------------------------------------------

                try:
                    headers, data_rows = dispatcher.run_processor(smeta_type=smeta_type, input_path=input_fpath)
                    if headers and data_rows is not None:
                        print(f"  ({client_session_id}) Получено строк данных: {len(data_rows)}")
                        collected_results.append((original_fname, headers, data_rows))
                        if common_headers is None: common_headers = headers
                        elif common_headers != headers: print("[WARN] Заголовки отличаются!")
                        # --- Обновляем processed count в статусе ПОСЛЕ УСПЕХА ---
                        processing_status[client_session_id]["processed"] += 1
                        # ------------------------------------------------------
                    else:
                        print(f"  [ОШИБКА] ({client_session_id}) Обработчик не вернул данные для {original_fname}")
                        processing_status[client_session_id]["error"] = f"Ошибка обработки {original_fname}" # Отмечаем ошибку в статусе
                        has_errors = True
                except Exception as e:
                    print(f"  [КРИТИЧЕСКАЯ ОШИБКА] ({client_session_id}) при обработке данных из {original_fname}: {e}")
                    processing_status[client_session_id]["error"] = f"Критическая ошибка при обработке {original_fname}"
                    traceback.print_exc(); has_errors = True

            # --- Анализ собранных данных ---
            if not collected_results:
                error_msg = "Во время обработки произошли ошибки, результаты не получены." if has_errors else "Не найдено данных для обработки."
                # Обновляем статус перед выбросом ошибки
                if client_session_id in processing_status:
                    processing_status[client_session_id]["status"] = "Ошибка"
                    processing_status[client_session_id]["error"] = error_msg
                raise ValueError(error_msg)

            # --- Обновляем статус: Подготовка итогового файла ---
            processing_status[client_session_id]["status"] = "Подготовка итогового файла..."
            # ---------------------------------------------------

            # --- Чтение референсных ширин ---
            reference_widths = None
            reference_file_to_read = None # Определяем какой файл читать
            if smeta_type == "Смета ру":
                reference_file_to_read = REFERENCE_SMETA_RU
            elif smeta_type.startswith("Турбосметчик-"): # <-- ИСПРАВЛЕНО: Проверяем начало строки
                 reference_file_to_read = REFERENCE_TURBOSMETCHIK

            if reference_file_to_read and os.path.exists(reference_file_to_read):
                print(f"({client_session_id}) Чтение референсных ширин из {os.path.basename(reference_file_to_read)}...") # <-- ИЗМЕНЕНО: Используем переменную
                try:
                     # Используем импортированный openpyxl
                     wb_ref = openpyxl.load_workbook(filename=reference_file_to_read, data_only=True, keep_vba=False) # <-- ИЗМЕНЕНО: Используем переменную
                     ws_ref = wb_ref.active; reference_widths = []
                     target_columns = ['A', 'B', 'C', 'D', 'E', 'F'] # Те же колонки A-F
                     for col_letter in target_columns: width = ws_ref.column_dimensions[col_letter].width if col_letter in ws_ref.column_dimensions else None; reference_widths.append(width if width is not None else 8.43) # Стандартная ширина 8.43, если не задана
                     print(f"  ({client_session_id}) Референсные ширины: {reference_widths}"); wb_ref.close()
                except Exception as e:
                    print(f"  [WARN] ({client_session_id}) Ошибка чтения реф. файла ({os.path.basename(reference_file_to_read)}): {e}.") # <-- ИЗМЕНЕНО: Используем переменную
                    reference_widths = None # Сбрасываем ширины при ошибке
            elif reference_file_to_read: # Если файл должен был быть, но его нет
                 print(f"  [WARN] ({client_session_id}) Реф. файл не найден: {reference_file_to_read}.")
            # --- Если reference_file_to_read is None (другой тип сметы), то reference_widths останется None ---

            # --- Создание и сохранение итогового файла ---
            final_wb = openpyxl.Workbook(); final_ws = final_wb.active # Используем openpyxl
            original_base_name = os.path.splitext(original_filename_unsafe)[0]

            # Очищаем базовое имя: разрешаем буквы (вкл. кириллицу), цифры, пробелы, _, -.
            # Заменяем другие потенциально проблемные символы на _.
            # Удаляем начальные/конечные пробелы и заменяем множественные пробелы одним.
            safe_base_name = re.sub(r'[^\w\s.-]+', '_', original_base_name, flags=re.UNICODE) # Оставляем буквы, цифры, _, пробел, -, . (Дефис в конце)
            safe_base_name = re.sub(r'\s+', ' ', safe_base_name).strip() # Убираем лишние пробелы
            # Дополнительно заменяем символы, опасные для файловых систем
            safe_base_name = re.sub(r'[\\\\/:*?"<>|]+', '_', safe_base_name)
            safe_base_name = safe_base_name.replace('..', '_') # Предотвращаем выход из директории

            if not safe_base_name: # Если имя стало пустым после очистки
                 safe_base_name = f"file_{uuid.uuid4().hex[:8]}" # Генерируем запасное имя

            output_filename = f"{safe_base_name}_processed.xlsx"
            final_ws.title = safe_base_name[:31] # Используем очищенное имя для названия листа

            # Используем наше очищенное имя. secure_filename больше не нужен здесь,
            # т.к. мы сами провели очистку.
            safe_output_filename = output_filename

            # Запасной вариант, если имя все равно оказалось некорректным (маловероятно)
            if not safe_output_filename or len(safe_output_filename) > 200: # Добавим проверку длины
                safe_output_filename = f"result_{client_session_id}.xlsx"

            output_file_path = os.path.join(app.config['RESULTS_FOLDER'], safe_output_filename)
            print(f"({client_session_id}) Подготовка итогового файла: {safe_output_filename}")

            if common_headers: final_ws.append(common_headers)
            print(f"({client_session_id}) Запись данных в итоговый файл...")
            for index, (original_fname, headers, data_rows) in enumerate(collected_results):
                if len(collected_results) > 1: # Добавляем разделитель, если файлов > 1
                    separator_row_idx = final_ws.max_row + 1; merge_range = f'A{separator_row_idx}:F{separator_row_idx}'
                    print(f"  ({client_session_id}) Доб. разделитель '{original_fname}' в строку {separator_row_idx}")
                    try:
                        final_ws.merge_cells(merge_range); cell = final_ws.cell(row=separator_row_idx, column=1)
                        cell.value = original_fname; cell.alignment = Alignment(horizontal='center', vertical='center'); cell.font = Font(bold=True)
                    except Exception as merge_err: print(f"  [WARN] ({client_session_id}) Ошибка merge разделителя: {merge_err}"); final_ws.cell(row=separator_row_idx, column=1).value = original_fname
                for row in data_rows: final_ws.append(row) # Добавляем данные
                print(f"  ({client_session_id}) Добавлено {len(data_rows)} строк из {original_fname}")

            # --- Обновляем статус: Форматирование ---
            processing_status[client_session_id]["status"] = "Применение форматирования..."
            # ---------------------------------------
            # --- ПРИМЕНЕНИЕ ШИРИН И ФОРМАТИРОВАНИЯ ---
            if reference_widths: # Применяем референсные ширины, если они были успешно прочитаны
                print(f"({client_session_id}) Применение референсных ширин...")
                apply_reference_widths(final_ws, reference_widths)
            else: # Иначе используем автоподбор
                print(f"({client_session_id}) Автоподбор ширины колонок...")
                auto_adjust_column_width(final_ws)
            # --- КОНЕЦ ИЗМЕНЕНИЙ В БЛОКЕ ПРИМЕНЕНИЯ ШИРИН ---
            apply_formatting(final_ws)

            # --- Обновляем статус: Сохранение ---
            processing_status[client_session_id]["status"] = "Сохранение файла..."
            # -----------------------------------
            final_wb.save(output_file_path); final_wb.close()
            print(f"({client_session_id}) Итоговый файл сохранен: {output_file_path}")

            # --- Успешный ответ ---
            final_message = "Обработка завершена."; download_url = url_for('download_file', filename=safe_output_filename)
            if has_errors: final_message += " Были ошибки при обработке некоторых файлов."
            if len(collected_results) < len(files_to_process_info): final_message += " Не все файлы из архива были успешно обработаны."

            # --- Обновляем статус: Готово ---
            processing_status[client_session_id]["status"] = "Готово"
            # -------------------------------

            return jsonify({"success": True, "message": final_message, "download_url": download_url, "download_filename": safe_output_filename })

        else: # Если файл не прошел allowed_file
             raise ValueError(f"Недопустимый тип файла: {original_filename_unsafe}.")

    except ValueError as ve: # Ловим ошибки типа файла, распаковки, отсутствия данных
         print(f"[ОШИБКА обработки] ({client_session_id}) {ve}")
         # --- Обновляем статус при ошибке ---
         if client_session_id in processing_status: processing_status[client_session_id]["status"] = "Ошибка"; processing_status[client_session_id]["error"] = str(ve)
         # ----------------------------------
         return jsonify({"success": False, "error": str(ve)}), 400
    except Exception as e: # Ловим все остальные ошибки
        print(f"[КРИТИЧЕСКАЯ ОШИБКА] ({client_session_id}) /upload: {e}")
        # --- Обновляем статус при ошибке ---
        if client_session_id in processing_status: processing_status[client_session_id]["status"] = "Критическая ошибка"; processing_status[client_session_id]["error"] = "Внутренняя ошибка сервера."
        # ----------------------------------
        traceback.print_exc()
        return jsonify({"success": False, "error": "Внутренняя ошибка сервера."}), 500
    finally:
        # Очистка временной папки загрузок
        if os.path.exists(upload_path):
            try: shutil.rmtree(upload_path); print(f"({client_session_id}) Очищена временная папка: {upload_path}")
            except Exception as clean_err: print(f"[WARN] ({client_session_id}) Не удалось очистить {upload_path}: {clean_err}")
        # Очистка статуса пока не реализована автоматически


# === Новый маршрут для получения прогресса ===
@app.route('/progress/<session_id>')
def get_progress(session_id):
    """Возвращает текущий статус обработки для данной сессии."""
    status = processing_status.get(session_id, {"status": "Не найдено", "processed": 0, "total": 0, "error": "Сессия не найдена или завершена."})
    # print(f"Запрос статуса для {session_id}: {status}") # Лог для отладки поллинга
    return jsonify(status)
# ==========================================

# --- Маршрут скачивания (ИЗМЕНЕН!) ---
@app.route('/download/<filename>')
def download_file(filename):
    # Мы предполагаем, что filename уже был очищен на этапе создания.
    # Убираем повторный вызов secure_filename, но проверяем на попытки выхода из папки.
    if '..' in filename or filename.startswith('/'):
        print(f"[ПРЕДУПРЕЖДЕНИЕ] Попытка доступа к небезопасному пути: {filename}")
        return "Недопустимое имя файла.", 400

    # Имя файла теперь используется как есть
    safe_filename = filename

    file_path = os.path.join(app.config['RESULTS_FOLDER'], safe_filename)
    if not os.path.exists(file_path): return "Файл не найден.", 404
    print(f"Отправка файла для скачивания: {safe_filename}")
    try: return send_from_directory(app.config['RESULTS_FOLDER'], safe_filename, as_attachment=True)
    except Exception as e: print(f"Ошибка при отправке файла {safe_filename}: {e}"); return "Ошибка при отправке файла.", 500

# --- Запуск приложения (без изменений) ---
if __name__ == '__main__':
    # debug=True включает автоперезагрузку при изменении кода и подробные ошибки в браузере
    # НЕ ИСПОЛЬЗОВАТЬ debug=True в production!
    app.run(host='0.0.0.0', port=5000, debug=True)
